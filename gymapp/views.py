from rest_framework import generics
from rest_framework.views import APIView
import razorpay
from django.conf import settings
from rest_framework.response import Response
from .models import Member, Attendance, Activity, Payment, MemberExercise
from rest_framework.decorators import api_view
from dateutil.relativedelta import relativedelta
from django.core.mail import send_mail
import calendar
from io import BytesIO
from django.core.mail import EmailMessage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from rest_framework import status
from .serializers import MemberSerializer, AttendanceSerializer, ActivitySerializer, PaymentSerializer
from datetime import date, timedelta
from datetime import datetime
import requests
from django.shortcuts import get_object_or_404
from django.http import FileResponse, JsonResponse
import base64
import random
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class MemberListView(generics.ListAPIView):
    queryset = Member.objects.all()
    serializer_class = MemberSerializer
class MemberCreateView(generics.CreateAPIView):
    queryset = Member.objects.all()
    serializer_class = MemberSerializer

    def perform_create(self, serializer):
        member = serializer.save()

        try:
            send_user_id_email(member)
        except BaseException as error:
            print(
                "Registration email failed:",
                type(error).__name__,
                str(error),
            )
class MemberDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Member.objects.all()
    serializer_class = MemberSerializer
    lookup_field = "id"

class LoginView(APIView):

    def post(self, request):

        user_id = request.data.get("user_id")
        password = request.data.get("password")

        try:
            member = Member.objects.get(user_id=user_id)

            if member.password == password:

                return Response({
                    "id": member.id,
                    "user_id": member.user_id,
                    "name": member.name,
                    "email": member.email,
                    "phone": member.phone,
                    "date_of_birth": member.date_of_birth,
                    "gender": member.gender,
                    "joined_date": member.joined_date,
                    "profile_image": member.profile_image.url,
                    "plan_type":member.plan_type,
                    "document_type":member.document_type,
                })

            return Response(
                {"message": "Invalid Password"},
                status=400
            )

        except Member.DoesNotExist:
            return Response(
                {"message": "User Not Found"},
                status=404
            )
class CheckInView(APIView):

    def post(self, request):

        user_id = request.data.get("user_id")

        try:
            member = Member.objects.get(
                user_id=user_id
            )

            attendance, created = Attendance.objects.get_or_create(
                member=member,
                date=date.today()
            )

            attendance.check_in = datetime.now().time()
            attendance.save()

            return Response({
                "message": "Check In Successful"
            })

        except Member.DoesNotExist:
            return Response(
                {"message": "Member Not Found"},
                status=404
            )
class CheckOutView(APIView):

    def post(self, request):

        user_id = request.data.get("user_id")

        try:

            member = Member.objects.get(
                user_id=user_id
            )

            attendance = Attendance.objects.get(
                member=member,
                date=date.today()
            )

            attendance.check_out = datetime.now().time()

            checkin = datetime.combine(
                date.today(),
                attendance.check_in
            )

            checkout = datetime.combine(
                date.today(),
                attendance.check_out
            )

            total = checkout - checkin

            attendance.total_hours = str(total)

            attendance.save()

            return Response({
                "message": "Check Out Successful",
                "hours": attendance.total_hours
            })

        except Attendance.DoesNotExist:
            return Response(
                {"message": "Check In First"},
                status=400
            )
class AttendanceHistoryView(APIView):

    def get(self, request, user_id):
        try:
            member = Member.objects.get(user_id=user_id)

            month = request.GET.get("month")
            year = request.GET.get("year")
            day = request.GET.get("day")

            attendance = Attendance.objects.filter(
                member=member
            )

            if month:
                attendance = attendance.filter(
                    date__month=int(month)
                )
            if day:
                attendance =attendance.filter(
                    date=day
                )    

            if year:
                attendance = attendance.filter(
                    date__year=int(year)
                )

            attendance = attendance.order_by("-date")

            serializer = AttendanceSerializer(
                attendance,
                many=True
            )

            return Response(serializer.data,)

        except Member.DoesNotExist:
            return Response(
                {"message": "Member Not Found"},
                status=404
            )
class AttendanceListView(generics.ListAPIView):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer

class MonthlyReportView(APIView):

    def get(self, request):
        month = request.GET.get("month")
        year = request.GET.get("year")

        today = date.today()

        month = int(month) if month else today.month
        year = int(year) if year else today.year

        if month == today.month and year == today.year:
            days_passed = today.day
        else:
            days_passed = calendar.monthrange(year, month)[1]

        members = Member.objects.all()
        data = []

        for member in members:
            attendance = Attendance.objects.filter(
                member=member,
                date__month=month,
                date__year=year
            )

            present_days = attendance.count()
            absent_days = days_passed - present_days

            attendance_percentage = round(
                (present_days / days_passed) * 100,
                2
            ) if days_passed > 0 else 0

            status = "Active" if attendance_percentage >= 50 else "Inactive"

            data.append({
                "user_id": member.user_id,
                "name": member.name,
                "present_days": present_days,
                "absent_days": absent_days,
                "attendance_percentage": attendance_percentage,
                "current_streak": 0,
                "status": status,
                "month": month,
                "year": year,
            })

        return Response(data)    
class ActivityListCreateView(
    generics.ListCreateAPIView
):
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer


class ActivityDetailView(
    generics.RetrieveUpdateDestroyAPIView
):
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer
    

class ActivityListView(generics.ListAPIView):
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer

@api_view(["POST"])
def create_order(request):

    amount = request.data.get("amount")

    client = razorpay.Client(
        auth=(
            settings.RAZORPAY_KEY_ID,
            settings.RAZORPAY_KEY_SECRET
        )
    )

    order = client.order.create({
        "amount": int(amount) * 100,
        "currency": "INR"
    })

    return Response({
        "order_id": order["id"],
        "amount": order["amount"],
        "key": settings.RAZORPAY_KEY_ID
    })



@api_view(["POST"])
def save_payment(request):
    user_id = request.data.get("user_id")
    plan = request.data.get("plan")
    amount = request.data.get("amount")
    validity = int(request.data.get("validity"))

    razorpay_payment_id = request.data.get("razorpay_payment_id")
    razorpay_order_id = request.data.get("razorpay_order_id")
    razorpay_signature = request.data.get("razorpay_signature")

    member = Member.objects.get(user_id=user_id)

    payment_mode = "Cash"

    if razorpay_payment_id and razorpay_payment_id != "dummy_payment":
        client = razorpay.Client(
            auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
        )

        payment_details = client.payment.fetch(razorpay_payment_id)
        payment_mode = payment_details.get("method", "Razorpay")
    else:
        payment_mode = request.data.get("payment_mode", "Cash")

    old_payment = Payment.objects.filter(
        member=member,
        status="Paid"
    ).order_by("-payment_date").first()

    new_expiry_date = date.today() + relativedelta(months=validity)

    if old_payment:
        remaining_days = (old_payment.expiry_date - date.today()).days

        if remaining_days > 5:
            return Response({
                "error": "You already have an active plan. Renewal allowed only before 5 days of expiry."
            }, status=400)

        old_payment.member_name = member.name
        old_payment.plan = plan
        old_payment.amount = amount
        old_payment.payment_mode = payment_mode
        old_payment.expiry_date = new_expiry_date
        old_payment.status = "Paid"
        old_payment.razorpay_order_id = razorpay_order_id
        old_payment.razorpay_payment_id = razorpay_payment_id
        old_payment.razorpay_signature = razorpay_signature
        old_payment.save()

        payment = old_payment

    else:
        payment = Payment.objects.create(
            member=member,
            member_name=member.name,
            plan=plan,
            amount=amount,
            payment_mode=payment_mode,
            expiry_date=new_expiry_date,
            status="Paid",
            razorpay_order_id=razorpay_order_id,
            razorpay_payment_id=razorpay_payment_id,
            razorpay_signature=razorpay_signature
        )
    send_payment_receipt_email(payment)
    return Response({
        "message": "Payment Saved Successfully",
        "plan": payment.plan,
        "amount": payment.amount,
        "payment_mode": payment.payment_mode,
        "expiry_date": payment.expiry_date,
        "status": payment.status
    })
@api_view(["GET"])
def my_membership(request, user_id):

    member = Member.objects.get(user_id=user_id)

    payment = Payment.objects.filter(
        member=member,
        status="Paid"
    ).order_by("-payment_date").first()

    if not payment:
        return Response({
            "status": "Inactive"
        })

    remaining_days = (
        payment.expiry_date - date.today()
    ).days
    can_renew = remaining_days <= 5
    return Response({
        "plan": payment.plan,
        "amount":payment.amount,
        "name":member.name,
        "status": payment.status if remaining_days >= 0 else "Expired",
        "start_date": payment.payment_date.date(),
        "expiry_date": payment.expiry_date,
        "remaining_days": remaining_days,
        "can_renew": can_renew,
        "payment_mode":payment.payment_mode


    })
@api_view(["GET"])
def membership_view(request):
    payments = Payment.objects.all().order_by("-payment_date")

    data = []

    for payment in payments:
        remaining_days = (payment.expiry_date - date.today()).days

        data.append({
            "id": payment.id,
            "member_name": payment.member.name,
            "plan": payment.plan,
            "amount": float(payment.amount),
            "status": payment.status if remaining_days >= 0 else "Expired",
            "payment_date": payment.payment_date.date(),
            "expiry_date": payment.expiry_date,
            "remaining_days": remaining_days,
            "payment_mode":payment.payment_mode,
            "payment_id":payment.razorpay_payment_id
        })

    return Response(data)

def send_expiry_warnings():
    target_date = date.today() + timedelta(days=5)

    payments = Payment.objects.filter(
        expiry_date=target_date,
        status="Paid",
        expiry_warning_sent=False
    )

    for payment in payments:
        member = payment.member

        subject = "Gym Membership Expiry Reminder"

        message = f"""
Hi {member.name},

Your {payment.plan} Plan at Infinity Wellness Hub will expire on {payment.expiry_date}.

Please renew your membership to continue your fitness journey.

Thank you,
Infinity Wellness Hub
"""

        send_mail(
            subject,
            message,
            "jasimr796@gmail.com",
            [member.email],
            fail_silently=False,
        )

        payment.expiry_warning_sent = True
        payment.save()


def send_payment_receipt_email(payment):
    try:
        print("Payment receipt function started")

        # Check required payment information
        if not payment.member.email:
            print("Member email is missing")
            return False

        buffer = BytesIO()

        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # Heading
        pdf.setFont("Helvetica-Bold", 18)
        pdf.drawCentredString(
            width / 2,
            height - 60,
            "Infinity Wellness Hub"
        )

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawCentredString(
            width / 2,
            height - 90,
            "Membership Payment Receipt"
        )

        # Receipt details
        y = height - 140

        payment_date = (
            payment.payment_date.strftime("%d/%m/%Y")
            if payment.payment_date
            else "Not available"
        )

        expiry_date = (
            payment.expiry_date.strftime("%d/%m/%Y")
            if payment.expiry_date
            else "Not available"
        )

        payment_mode = getattr(
            payment,
            "payment_mode",
            None
        ) or "Online"

        details = [
            ("Member Name", payment.member.name),
            ("User ID", payment.member.user_id),
            ("Plan", payment.plan),
            ("Amount", f"Rs. {payment.amount}"),
            ("Payment Method", payment_mode),
            ("Payment Date", payment_date),
            ("Expiry Date", expiry_date),
            ("Status", payment.status),
        ]

        for label, value in details:
            pdf.setFont("Helvetica-Bold", 11)
            pdf.drawString(80, y, f"{label}:")

            pdf.setFont("Helvetica", 11)
            pdf.drawString(230, y, str(value))

            y -= 25

        pdf.setFont("Helvetica", 10)
        pdf.drawCentredString(
            width / 2,
            60,
            "Thank you for choosing Infinity Wellness Hub."
        )

        pdf.save()

        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        encoded_pdf = base64.b64encode(
            pdf_bytes
        ).decode("utf-8")

        url = "https://api.brevo.com/v3/smtp/email"

        headers = {
            "accept": "application/json",
            "api-key": settings.BREVO_API_KEY,
            "content-type": "application/json",
        }

        data = {
            "sender": {
                "name": "Infinity Wellness Hub",
                "email": settings.DEFAULT_FROM_EMAIL,
            },

            "to": [
                {
                    "email": payment.member.email,
                    "name": payment.member.name,
                }
            ],

            "subject": (
                "Infinity Wellness Hub - Payment Receipt"
            ),

            "htmlContent": f"""
            <html>
                <body style="
                    font-family: Arial, sans-serif;
                    color: #222222;
                ">

                    <h2 style="color: #ff6600;">
                        Payment Successful
                    </h2>

                    <p>
                        Dear {payment.member.name},
                    </p>

                    <p>
                        Your membership payment has been
                        recorded successfully.
                    </p>

                    <h3>Payment Details</h3>

                    <p>
                        <strong>Member ID:</strong>
                        {payment.member.user_id}
                        <br>

                        <strong>Plan:</strong>
                        {payment.plan}
                        <br>

                        <strong>Amount:</strong>
                        Rs. {payment.amount}
                        <br>

                        <strong>Payment Method:</strong>
                        {payment_mode}
                        <br>

                        <strong>Payment Date:</strong>
                        {payment_date}
                        <br>

                        <strong>Expiry Date:</strong>
                        {expiry_date}
                        <br>

                        <strong>Status:</strong>
                        {payment.status}
                    </p>

                    <p>
                        Your payment receipt is attached
                        to this email.
                    </p>

                    <p>
                        Thank you for choosing
                        Infinity Wellness Hub.
                    </p>

                    <p>
                        Regards,
                        <br>
                        <strong>
                            Infinity Wellness Hub
                        </strong>
                    </p>

                </body>
            </html>
            """,

            "attachment": [
                {
                    "content": encoded_pdf,
                    "name": (
                        f"{payment.member.user_id}"
                        "-payment-receipt.pdf"
                    ),
                }
            ],
        }

        response = requests.post(
            url,
            headers=headers,
            json=data,
            timeout=30,
        )

        print(
            "Payment email status:",
            response.status_code
        )

        print(
            "Payment email response:",
            response.text
        )

        if response.status_code in [200, 201, 202]:
            print("Payment receipt email sent successfully")
            return True

        print("Payment receipt email failed")
        return False

    except Exception as error:
        print(
            "Payment receipt email error:",
            str(error)
        )
        return False



class AssignWorkoutView(APIView):

    def post(self, request):

       
        member_id = request.data.get("member_id")
        member_day = request.data.get("member_day")
        workout_day = request.data.get("workout_day")

        try:
            
            member = Member.objects.get(user_id=member_id)
            if workout_day and workout_day.strip() != "":
                duplicate_workout = MemberExercise.objects.filter(
                    member=member,
                    workout_day=workout_day
                ).exclude(
                    member_day=member_day
                ).exists()

                if duplicate_workout:
                    return Response({
                        "error": f"{workout_day} is already assigned for this member in this week."
                    }, status=400)

            MemberExercise.objects.update_or_create(
                member=member,
                member_day=member_day,
                defaults={
                    "trainer": "Trainer",
                    "workout_day": workout_day
                }
            )

            return Response({
                "message": "Workout assigned successfully."
            })

        except Member.DoesNotExist:
            return Response(
                {"error": "Member or Trainer not found"},
                status=status.HTTP_404_NOT_FOUND
            )



class MemberWorkoutView(APIView):

    def get(self, request, user_id):
        today = date.today().strftime("%A")

        try:
            plan = MemberExercise.objects.get(
                member__user_id=user_id,
                member_day=today
            )

            activities = Activity.objects.filter(
                workout_day=plan.workout_day
            )

            data = []

            for item in activities:
                data.append({
                    "exercise_name": item.exercise_name,
                    "sets": item.sets,
                    "description": item.description,
                    "image": request.build_absolute_uri(item.image.url) if item.image else None,
                })

            return Response({
                "member_day": today,
                "workout_day": plan.workout_day,
                "activities": data
            })

        except MemberExercise.DoesNotExist:
            return Response({
                "member_day": today,
                "workout_day": "",
                "activities": []
            })
class AllMemberWorkoutTableView(APIView):

    def get(self, request):
        members = Member.objects.all()

        data = []

        for member in members:
            plans = MemberExercise.objects.filter(member=member)

            member_data = {
                "user_id": member.user_id,
                "name": member.name,
                "Monday": "",
                "Tuesday": "",
                "Wednesday": "",
                "Thursday": "",
                "Friday": "",
                "Saturday": "",
            }

            for plan in plans:
                member_data[plan.member_day] = plan.workout_day

            data.append(member_data)

        return Response(data)



@api_view(["GET"])
def dashboard(request):
    user_id = request.GET.get("user_id")

    if not user_id:
        return Response(
            {"detail": "user_id is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        member = Member.objects.get(user_id=user_id)
    except Member.DoesNotExist:
        return Response(
            {"detail": "Member not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    latest_payment = (
        Payment.objects
        .filter(member=member, status="Paid")
        .order_by("-payment_date")
        .first()
    )

    membership = None

    if latest_payment:
        remaining_days = (
            latest_payment.expiry_date - date.today()
        ).days

        membership = {
            "plan": latest_payment.plan,
            "amount": latest_payment.amount,
            "status": "Active" if remaining_days >= 0 else "Expired",
            "start_date": latest_payment.payment_date.date(),
            "expiry_date": latest_payment.expiry_date,
            "remaining_days": max(remaining_days, 0),
        }

    attendance_records = Attendance.objects.filter(member=member)

    attendance = {
        "present_days": attendance_records.filter(
            status="Present"
        ).count(),
        "absent_days": attendance_records.filter(
            status="Absent"
        ).count(),
    }

    today_name = date.today().strftime("%A")

    today_workouts = Activity.objects.filter(day=today_name)

    return Response({
        "member": MemberSerializer(member).data,
        "membership": membership,
        "attendance": attendance,
        "today_workout": ActivitySerializer(
            today_workouts,
            many=True
        ).data,
    })


def send_user_id_email(member):
    url = "https://api.brevo.com/v3/smtp/email"

    headers = {
        "accept": "application/json",
        "api-key": settings.BREVO_API_KEY,
        "content-type": "application/json",
    }

    data = {
        "sender": {
            "name": "Infinity Wellness Hub",
            "email": settings.DEFAULT_FROM_EMAIL,
        },
        "to": [
            {
                "email": member.email,
                "name": member.name,
            }
        ],
        "subject": "Registration Successful",
        "textContent": f"""
Welcome {member.name}

Welcome to Infinity Wellness Hub!

Your registration has been completed successfully. We are excited to have you as a member of our fitness family.

Your Membership Details:
Member Name : {member.name}
Member ID : {member.user_id}
Registration Date : {member.joined_date}
Member ID: {member.user_id}

Please keep your Member ID safe, as it will be required for attendance, membership renewals, and other gym services.

You can now log in to the Infinity Wellness Hub mobile application using your registered email and password.

If you have any questions or need assistance, feel free to contact us.

Thank you for choosing Infinity Wellness Hub.

Stay Healthy. Stay Strong.

Thank you for choosing Infinity Wellness Hub.
""",
    }

    response = requests.post(url, headers=headers, json=data)
    print(response.status_code)
    print(response.text)

def member_payment_pdf(request, user_id):
    member = get_object_or_404(
        Member,
        user_id=user_id
    )

    payment = (
        Payment.objects
        .filter(member=member)
        .order_by("-payment_date")
        .first()
    )

    if not payment:
        return JsonResponse(
            {"message": "No payment found for this member"},
            status=404,
        )

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    pdf.setTitle(
        f"{member.name} Payment Report"
    )

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawCentredString(
        width / 2,
        height - 60,
        "Infinity Wellness Hub"
    )

    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(
        width / 2,
        height - 90,
        "Member Payment Report"
    )

    pdf.line(
        60,
        height - 110,
        width - 60,
        height - 110
    )

    y = height - 155

    details = [
        ("Member Name", member.name),
        ("User ID", member.user_id),
        ("Email", member.email),
        ("Phone", member.phone),
        ("Plan", payment.plan),
        ("Amount", f"Rs. {payment.amount}"),
        (
            "Payment Mode",
            payment.payment_mode or "Online"
        ),
        (
            "Payment Date",
            payment.payment_date.strftime("%d/%m/%Y")
        ),
        (
            "Expiry Date",
            payment.expiry_date.strftime("%d/%m/%Y")
        ),
        ("Status", payment.status),
    ]

    for label, value in details:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(70, y, f"{label}:")

        pdf.setFont("Helvetica", 11)
        pdf.drawString(220, y, str(value))

        y -= 28

    pdf.line(
        60,
        y - 5,
        width - 60,
        y - 5
    )

    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(
        width / 2,
        60,
        "Thank you for choosing Infinity Wellness Hub."
    )

    pdf.save()
    buffer.seek(0)

    filename = (
        f"{member.user_id}-payment-report.pdf"
    )

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type="application/pdf",
    )
class MemberWorkoutAllView(APIView):

    def get(self, request, user_id):
        try:
            member = Member.objects.get(user_id=user_id)
        except Member.DoesNotExist:
            return Response(
                {"error": "Member not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        plans = MemberExercise.objects.filter(member=member)

        member_data = {
            "user_id": member.user_id,
            "name": member.name,
            "Monday": "",
            "Tuesday": "",
            "Wednesday": "",
            "Thursday": "",
            "Friday": "",
            "Saturday": "",
        }

        for plan in plans:
            if plan.member_day in member_data:
                member_data[plan.member_day] = plan.workout_day

        return Response(member_data, status=status.HTTP_200_OK)



@api_view(["POST"])
def forgot_password(request):
    email = request.data.get("email", "").strip().lower()

    if not email:
        return Response(
            {"message": "Email required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    member = Member.objects.filter(email__iexact=email).first()

    if not member:
        return Response(
            {"message": "Email not found"},
            status=status.HTTP_404_NOT_FOUND,
        )

    try:
        otp = str(random.randint(100000, 999999))

        member.otp = otp
        member.otp_created_at = timezone.now()
        member.save(update_fields=["otp", "otp_created_at"])

        brevo_api_key = getattr(settings, "BREVO_API_KEY", None)
        sender_email = getattr(settings, "DEFAULT_FROM_EMAIL", None)

        if not brevo_api_key:
            return Response(
                {"message": "BREVO_API_KEY is not configured"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        if not sender_email:
            return Response(
                {"message": "SENDER_EMAIL is not configured"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        headers = {
            "accept": "application/json",
            "api-key": brevo_api_key,
            "content-type": "application/json",
        }

        html = f"""
        <h2>Infinity Wellness Hub</h2>
        <p>Hello {member.name}</p>
        <p>Your password reset OTP is:</p>
        <h1 style="color:#ff7a00;">{otp}</h1>
        <p>This OTP is valid for 10 minutes.</p>
        """

        payload = {
            "sender": {
                "name": "Infinity Wellness Hub",
                "email": sender_email,
            },
            "to": [
                {
                    "email": member.email,
                    "name": member.name,
                }
            ],
            "subject": "Password Reset OTP",
            "htmlContent": html,
        }

        brevo_response = requests.post(
            "https://api.brevo.com/v3/smtp/email",
            headers=headers,
            json=payload,
            timeout=20,
        )

        print("BREVO STATUS:", brevo_response.status_code)
        print("BREVO RESPONSE:", brevo_response.text)

        if brevo_response.status_code not in [200, 201, 202]:
            return Response(
                {
                    "message": "OTP created, but email sending failed",
                    "email_service_error": brevo_response.text,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {"message": "OTP Sent Successfully"},
            status=status.HTTP_200_OK,
        )

    except Exception as error:
        print("FORGOT PASSWORD ERROR:", repr(error))

        return Response(
            {
                "message": "Server error while sending OTP",
                "error": str(error),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
@api_view(["POST"])
def verify_otp(request):

    email = request.data.get("email")
    otp = request.data.get("otp")

    try:
        member = Member.objects.get(email=email)

    except Member.DoesNotExist:
        return Response(
            {"message": "Member Not Found"},
            status=404
        )

    if member.otp != otp:
        return Response(
            {"message": "Invalid OTP"},
            status=400
        )

    member.otp = None
    member.save()

    return Response({
        "message": "OTP Verified",
        "member_id": member.id
    })

def export_member_fee_report(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Member Fee Report"

    
    worksheet.merge_cells("A1:F1")

    title_cell = worksheet["A1"]
    title_cell.value = "Infinity Wellness Hub - Member Fee Report"
    title_cell.font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )
    title_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="FF6600"
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 30

    
    headers = [
        "Member ID",
        "Member Name",
        "Total Amount Paid",
        "Current Membership",
        "Current Amount",
        "Total Payments",
    ]

    header_row = 3

    for column_number, heading in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=heading
        )

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="222222"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    paid_members = Member.objects.all().order_by("name")

    data_row = 4

    for member in paid_members:
    
        payments = Payment.objects.filter(
            member=member,
            status="Paid"
        ).order_by("-payment_date")

    
        if not payments.exists():
            continue

        summary = payments.aggregate(
            total_amount=Sum("amount"),
            total_payments=Count("id")
        )

        latest_payment = payments.first()

        worksheet.cell(
            row=data_row,
            column=1,
            value=member.user_id
        )

        worksheet.cell(
            row=data_row,
            column=2,
            value=member.name
        )

        worksheet.cell(
            row=data_row,
            column=3,
            value=float(summary["total_amount"] or 0)
        )

        worksheet.cell(
            row=data_row,
            column=4,
            value=latest_payment.plan
        )

        worksheet.cell(
            row=data_row,
            column=5,
            value=float(latest_payment.amount or 0)
        )

        worksheet.cell(
            row=data_row,
            column=6,
            value=summary["total_payments"] or 0
        )

        data_row += 1

    
    for row in range(4, data_row):
        worksheet.cell(row=row, column=3).number_format = '₹#,##0.00'
        worksheet.cell(row=row, column=5).number_format = '₹#,##0.00'

    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for row in worksheet.iter_rows(
        min_row=3,
        max_row=max(data_row - 1, 3),
        min_col=1,
        max_col=6
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    
    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 25
    worksheet.column_dimensions["C"].width = 22
    worksheet.column_dimensions["D"].width = 24
    worksheet.column_dimensions["E"].width = 20
    worksheet.column_dimensions["F"].width = 18

    # Freeze table heading
    worksheet.freeze_panes = "A4"

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="member_fee_report.xlsx"'
    )

    workbook.save(response)

    return response
def format_time(value):
    
    if not value:
        return None

    return value.strftime("%I:%M:%S %p")


def calculate_total_hours(check_in, check_out):
    
    if not check_in or not check_out:
        return ""

    today = timezone.localdate()

    check_in_datetime = timezone.datetime.combine(
        today,
        check_in
    )

    check_out_datetime = timezone.datetime.combine(
        today,
        check_out
    )

    
    if check_out_datetime < check_in_datetime:
        check_out_datetime += timezone.timedelta(days=1)

    duration = check_out_datetime - check_in_datetime

    total_seconds = int(duration.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60

    return f"{hours}:{minutes}"


@api_view(["POST"])
def biometric_scan(request):
    try:
        employee_id = request.data.get("employee_id")

        device_id = request.data.get(
            "device_id",
            "DEMO-BIO-001"
        )

        verification_type = request.data.get(
            "verification_type",
            "Fingerprint"
        )

        scan_status = request.data.get(
            "status",
            "verified"
        )

        if not employee_id:
            return Response(
                {
                    "success": False,
                    "message": "Member ID is required."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if str(scan_status).lower() != "verified":
            return Response(
                {
                    "success": False,
                    "message": "Biometric verification failed."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            member = Member.objects.get(
                user_id=employee_id
            )

        except Member.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Member not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        today = timezone.localdate()

        local_time = timezone.localtime(
            timezone.now()
        ).time().replace(microsecond=0)

        attendance, created = Attendance.objects.get_or_create(
            member=member,
            date=today,
            defaults={
                "check_in": local_time,
                "status": "Present",
                "attendance_method": verification_type,
                "device_id": device_id,
            }
        )

        # -----------------------------
        # First Scan -> Check In
        # -----------------------------
        if created:
            action = "check_in"
            message = "Biometric check-in successful."

        elif attendance.check_in is None:

            attendance.check_in = local_time
            attendance.status = "Present"
            attendance.attendance_method = verification_type
            attendance.device_id = device_id

            attendance.save(
                update_fields=[
                    "check_in",
                    "status",
                    "attendance_method",
                    "device_id",
                ]
            )

            action = "check_in"
            message = "Biometric check-in successful."

        # -----------------------------
        # Second Scan -> Check Out
        # -----------------------------
        elif attendance.check_out is None:

            check_in_datetime = datetime.combine(
                attendance.date,
                attendance.check_in
            )

            current_datetime = timezone.localtime(
                timezone.now()
            ).replace(tzinfo=None)

            time_difference = current_datetime - check_in_datetime

            # Check if 30 minutes completed
            if time_difference < timedelta(minutes=30):

                remaining = timedelta(minutes=30) - time_difference

                remaining_minutes = int(
                    remaining.total_seconds() // 60
                )

                return Response(
                    {
                        "success": False,
                        "message": (
                            "Check-in is already recorded. "
                            f"Check-out is allowed after "
                            f"{remaining_minutes} minute(s)."
                        ),
                        "action": "check_in_already_done",
                        "member": {
                            "user_id": member.user_id,
                            "name": member.name,
                        },
                        "attendance": {
                            "date": str(attendance.date),
                            "check_in": format_time(
                                attendance.check_in
                            ),
                            "check_out": None,
                            "total_hours": attendance.total_hours,
                            "method": attendance.attendance_method,
                            "device_id": attendance.device_id,
                        }
                    },
                    status=status.HTTP_200_OK
                )

            attendance.check_out = local_time
            attendance.attendance_method = verification_type
            attendance.device_id = device_id

            attendance.total_hours = calculate_total_hours(
                attendance.check_in,
                attendance.check_out
            )

            attendance.save(
                update_fields=[
                    "check_out",
                    "total_hours",
                    "attendance_method",
                    "device_id",
                ]
            )

            action = "check_out"
            message = "Biometric check-out successful."

        # -----------------------------
        # Third Scan
        # -----------------------------
        else:

            return Response(
                {
                    "success": False,
                    "message": (
                        "Today's check-in and check-out "
                        "are already completed."
                    ),
                    "action": "completed",
                    "member": {
                        "user_id": member.user_id,
                        "name": member.name,
                    },
                    "attendance": {
                        "date": str(attendance.date),
                        "check_in": format_time(
                            attendance.check_in
                        ),
                        "check_out": format_time(
                            attendance.check_out
                        ),
                        "total_hours": attendance.total_hours,
                        "method": attendance.attendance_method,
                        "device_id": attendance.device_id,
                    }
                },
                status=status.HTTP_200_OK
            )

        attendance.refresh_from_db()

        return Response(
            {
                "success": True,
                "message": message,
                "action": action,
                "member": {
                    "user_id": member.user_id,
                    "name": member.name,
                },
                "attendance": {
                    "date": str(attendance.date),
                    "check_in": format_time(
                        attendance.check_in
                    ),
                    "check_out": format_time(
                        attendance.check_out
                    ),
                    "total_hours": attendance.total_hours,
                    "method": attendance.attendance_method,
                    "device_id": attendance.device_id,
                }
            },
            status=status.HTTP_200_OK
        )

    except Exception as error:

        print(
            "BIOMETRIC ATTENDANCE ERROR:",
            repr(error)
        )

        return Response(
            {
                "success": False,
                "message": "Biometric attendance server error.",
                "error": str(error),
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )